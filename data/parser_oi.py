"""Parse weekly open interest Excel files."""

import io
import re
import openpyxl
from datetime import date
from typing import Optional
from models import ParticipantOI
import config


def parse_oi_excel(
    content: bytes,
    target_products: Optional[list[str]] = None,
) -> list[ParticipantOI]:
    """Parse a weekly open interest Excel file.

    Returns list of ParticipantOI records with long_volume and short_volume
    consolidated per (product, contract_month, participant_id).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    report_date = _extract_report_date(ws)

    if target_products is None:
        target_products = config.TARGET_PRODUCTS

    section_headers = _find_section_headers(ws)

    results = []
    for product_code, header_row in section_headers.items():
        if product_code not in target_products:
            continue

        data_start = header_row + config.OI_DATA_OFFSET

        # Parse near month (left half, columns A-H)
        near_records = _parse_oi_half(
            ws, data_start, product_code, report_date,
            col_offsets=config.OI_NEAR_COLUMNS,
        )
        results.extend(near_records)

        # Parse far month (right half, columns K-R)
        far_records = _parse_oi_half(
            ws, data_start, product_code, report_date,
            col_offsets=config.OI_FAR_COLUMNS,
        )
        results.extend(far_records)

    wb.close()
    return _consolidate_long_short(results)


def _find_section_headers(ws) -> dict[str, int]:
    """Scan column A for product section header rows.

    Returns: {product_code: header_row_number}
    """
    headers = {}
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is None or isinstance(val, (int, float)):
            continue
        text = str(val)
        text_lower = text.lower()
        # Check patterns (order matters: mini before generic 225)
        if "mini" in text_lower and "225" in text:
            headers["NK225MF"] = row_idx
        elif "225" in text and "mini" not in text_lower:
            headers["NK225F"] = row_idx
        elif "topix" in text_lower:
            headers["TOPIXF"] = row_idx
    return headers


def _extract_report_date(ws) -> date:
    """Extract the report date from header rows.

    Row 2: '（ 2026年01月30日現在 ）' -> date(2026, 1, 30)
    """
    for row_idx in [2, 3, 1]:
        val = ws.cell(row=row_idx, column=1).value
        if val:
            digits = re.findall(r'\d+', str(val))
            if len(digits) >= 3:
                y, m, d = int(digits[0]), int(digits[1]), int(digits[2])
                if 2000 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
                    return date(y, m, d)
    raise ValueError("Could not extract report date from OI Excel file")


def _parse_contract_month_yymm(text: str) -> str:
    """Parse '2026年03月限月' -> '2603'."""
    digits = re.findall(r'\d+', str(text))
    if len(digits) >= 2:
        year = digits[0]
        month = digits[1].zfill(2)
        return year[2:] + month
    return ""


def _parse_oi_half(
    ws,
    data_start: int,
    product_code: str,
    report_date: date,
    col_offsets: dict,
) -> list[ParticipantOI]:
    """Parse one half (near or far) of an OI section."""
    # Read contract month from first data row
    cm_cell = ws.cell(row=data_start, column=col_offsets["contract_month"]).value
    if not cm_cell:
        return []
    contract_month = _parse_contract_month_yymm(str(cm_cell))
    if not contract_month:
        return []

    records = []

    for i in range(config.OI_ROWS_PER_SECTION):
        row = data_start + i
        rank = ws.cell(row=row, column=col_offsets["rank"]).value
        if rank is None:
            break

        # Long side
        long_pid = ws.cell(row=row, column=col_offsets["long_pid"]).value
        long_name = ws.cell(row=row, column=col_offsets["long_name_jp"]).value
        long_vol = ws.cell(row=row, column=col_offsets["long_volume"]).value

        if long_pid:
            records.append(ParticipantOI(
                report_date=report_date,
                product=product_code,
                contract_month=contract_month,
                participant_id=str(long_pid),
                participant_name_jp=str(long_name or ""),
                long_volume=float(long_vol) if long_vol else 0.0,
                short_volume=None,
            ))

        # Short side
        short_pid = ws.cell(row=row, column=col_offsets["short_pid"]).value
        short_name = ws.cell(row=row, column=col_offsets["short_name_jp"]).value
        short_vol = ws.cell(row=row, column=col_offsets["short_volume"]).value

        if short_pid:
            records.append(ParticipantOI(
                report_date=report_date,
                product=product_code,
                contract_month=contract_month,
                participant_id=str(short_pid),
                participant_name_jp=str(short_name or ""),
                long_volume=None,
                short_volume=float(short_vol) if short_vol else 0.0,
            ))

    return records


def _consolidate_long_short(records: list[ParticipantOI]) -> list[ParticipantOI]:
    """Merge records so each (product, contract_month, participant_id) has
    both long_volume and short_volume in a single record."""
    key_map: dict[tuple, ParticipantOI] = {}
    for rec in records:
        key = (rec.product, rec.contract_month, rec.participant_id)
        if key in key_map:
            existing = key_map[key]
            if rec.long_volume is not None:
                existing.long_volume = rec.long_volume
            if rec.short_volume is not None:
                existing.short_volume = rec.short_volume
        else:
            key_map[key] = rec
    return list(key_map.values())
