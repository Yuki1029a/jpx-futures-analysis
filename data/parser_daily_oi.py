"""Parse daily OI balance Excel (open_interest_e.xlsx Attachment1).

Sheet structure (Attachment1 = NK225 Options per strike):
  Row 2: Report date (datetime)
  Row 5: Section headers -- PUT left, CALL right
  Row 6: Column headers
  Row 7+: Data rows

  PUT cols: A=contract code, B=volume, C=current_oi, D=change, E=prev_oi
  CALL cols: G=contract code, H=volume, I=current_oi, J=change, K=prev_oi

  Contract code format: "NIKKEI 225 P2603-38000"
  PUT and CALL sides are NOT row-aligned.
  Blocks separated by "Total for Contract Month" rows.
"""
from __future__ import annotations

import io
import re
import logging
from datetime import date, datetime

import openpyxl

from models import DailyOIBalance, DailyFuturesOI

logger = logging.getLogger(__name__)

_CONTRACT_RE = re.compile(r'NIKKEI 225 ([PC])(\d{4})-(\d+)')

# Contract month: "2026年03月限" → extract year + month
_CM_RE = re.compile(r'(\d{4})\D+(\d{1,2})\D')

# Product detection by Unicode content in cell text
# Map: substring to check → product code
_FUTURES_PRODUCTS_LEFT = {
    '\u65e5\u7d4c225': 'NK225F',    # 日経225 (in A column)
    'TOPIX': 'TOPIXF',
}
_FUTURES_PRODUCTS_RIGHT = {
    '\u65e5\u7d4c225mini': 'NK225MF',   # 日経225mini (in H column)
    '\u65e5\u7d4c225\u30de\u30a4\u30af\u30ed': 'NK225MicroF',  # 日経225マイクロ
    '\u30df\u30cbTOPIX': 'MiniTOPIXF',
    'JPX\u65e5\u7d4c400': 'JPX400F',
}


def parse_daily_oi_excel(content: bytes) -> list[DailyOIBalance]:
    """Parse Attachment1 sheet of daily OI balance Excel.

    Returns all PUT and CALL records across all contract months.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Use second sheet (Attachment1); index-based since JP version has garbled names
    if len(wb.worksheets) < 2:
        logger.warning("Daily OI Excel has fewer than 2 sheets")
        wb.close()
        return []

    ws = wb.worksheets[1]
    report_date = _extract_date(ws)
    results: list[DailyOIBalance] = []

    for row_idx in range(7, ws.max_row + 1):
        # PUT side (col A=1)
        put_code = ws.cell(row=row_idx, column=1).value
        if put_code:
            m = _CONTRACT_RE.search(str(put_code))
            if m and m.group(1) == "P":
                results.append(DailyOIBalance(
                    report_date=report_date,
                    contract_month=m.group(2),
                    option_type="PUT",
                    strike_price=int(m.group(3)),
                    trading_volume=_safe_int(ws.cell(row=row_idx, column=2).value),
                    current_oi=_safe_int(ws.cell(row=row_idx, column=3).value),
                    net_change=_safe_int(ws.cell(row=row_idx, column=4).value),
                    previous_oi=_safe_int(ws.cell(row=row_idx, column=5).value),
                ))

        # CALL side (col G=7)
        call_code = ws.cell(row=row_idx, column=7).value
        if call_code:
            m = _CONTRACT_RE.search(str(call_code))
            if m and m.group(1) == "C":
                results.append(DailyOIBalance(
                    report_date=report_date,
                    contract_month=m.group(2),
                    option_type="CALL",
                    strike_price=int(m.group(3)),
                    trading_volume=_safe_int(ws.cell(row=row_idx, column=8).value),
                    current_oi=_safe_int(ws.cell(row=row_idx, column=9).value),
                    net_change=_safe_int(ws.cell(row=row_idx, column=10).value),
                    previous_oi=_safe_int(ws.cell(row=row_idx, column=11).value),
                ))

    wb.close()
    return results


def _extract_date(ws) -> date:
    """Extract report date from row 2, column A."""
    val = ws.cell(row=2, column=1).value
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Fallback: extract digits
    digits = re.findall(r'\d+', str(val))
    if len(digits) >= 3:
        return date(int(digits[0]), int(digits[1]), int(digits[2]))
    raise ValueError(f"Cannot extract date from daily OI Excel: {val}")


def parse_daily_futures_oi_excel(content: bytes) -> list[DailyFuturesOI]:
    """Parse Sheet 0 (デリバティブ建玉残高) for futures OI balance.

    Sheet 0 layout:
      Row 2: report date
      Sections separated by section headers (e.g. ＜OSE 指数先物取引＞)
      Left side (cols A-F): product in A, contract months in B, data in C-F
      Right side (cols H-M): product in H, contract months in I, data in J-M

    Each product block: product name row → contract month rows → 合計 row.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    report_date = _extract_date(ws)
    results: list[DailyFuturesOI] = []

    current_product_left: str | None = None
    current_product_right: str | None = None

    for row_idx in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        h_val = ws.cell(row=row_idx, column=8).value

        # Detect product names
        if a_val and isinstance(a_val, str):
            matched = _match_product(a_val, _FUTURES_PRODUCTS_LEFT)
            if matched:
                current_product_left = matched
            elif '\u5408\u8a08' in a_val:  # 合計
                current_product_left = None

        if h_val and isinstance(h_val, str):
            matched = _match_product(h_val, _FUTURES_PRODUCTS_RIGHT)
            if matched:
                current_product_right = matched
            elif '\u5408\u8a08' in h_val:  # 合計
                current_product_right = None

        # Parse left side data (B=contract_month, C=volume, D=current_oi, E=change, F=prev_oi)
        if current_product_left:
            cm = _parse_contract_month(ws.cell(row=row_idx, column=2).value)
            if cm:
                results.append(DailyFuturesOI(
                    report_date=report_date,
                    product=current_product_left,
                    contract_month=cm,
                    trading_volume=_safe_int(ws.cell(row=row_idx, column=3).value),
                    current_oi=_safe_int(ws.cell(row=row_idx, column=4).value),
                    net_change=_safe_int(ws.cell(row=row_idx, column=5).value),
                    previous_oi=_safe_int(ws.cell(row=row_idx, column=6).value),
                ))

        # Parse right side data (I=contract_month, J=volume, K=current_oi, L=change, M=prev_oi)
        if current_product_right:
            cm = _parse_contract_month(ws.cell(row=row_idx, column=9).value)
            if cm:
                results.append(DailyFuturesOI(
                    report_date=report_date,
                    product=current_product_right,
                    contract_month=cm,
                    trading_volume=_safe_int(ws.cell(row=row_idx, column=10).value),
                    current_oi=_safe_int(ws.cell(row=row_idx, column=11).value),
                    net_change=_safe_int(ws.cell(row=row_idx, column=12).value),
                    previous_oi=_safe_int(ws.cell(row=row_idx, column=13).value),
                ))

    wb.close()
    return results


def _match_product(cell_text: str, product_map: dict[str, str]) -> str | None:
    """Match cell text against product name patterns."""
    for pattern, code in product_map.items():
        if pattern in cell_text:
            return code
    return None


def _parse_contract_month(val) -> str | None:
    """Parse contract month cell: '2026年03月限' → '2603'."""
    if val is None:
        return None
    s = str(val)
    if '\u5408\u8a08' in s:  # 合計
        return None
    m = _CM_RE.search(s)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        return f"{year % 100:02d}{month:02d}"
    return None


def _safe_int(val) -> int:
    """Convert cell value to int, defaulting to 0."""
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0
