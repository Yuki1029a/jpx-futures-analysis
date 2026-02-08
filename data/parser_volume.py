"""Parse daily trading volume Excel files."""

import io
import re
import openpyxl
from datetime import datetime
from typing import Optional
from models import ParticipantVolume, OptionParticipantVolume
import config


def parse_volume_excel(
    content: bytes,
    target_products: Optional[list[str]] = None,
) -> list[ParticipantVolume]:
    """Parse a single daily volume Excel file (WholeDay or Night).

    Returns list of ParticipantVolume records.
    volume_day or volume_night is set depending on session; the other is 0.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Extract trade date from cell C5 (e.g., "20260206")
    trade_date_str = str(ws.cell(row=5, column=3).value).strip()
    trade_date = datetime.strptime(trade_date_str, "%Y%m%d").date()

    # Detect session type from row 1 or row 2 text
    header_text = str(ws.cell(row=2, column=1).value or "")
    is_night = "Night" in header_text

    if target_products is None:
        target_products = config.TARGET_PRODUCTS

    results = []
    for row_idx in range(config.VOLUME_DATA_START_ROW, ws.max_row + 1):
        product = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["product"]).value
        if product is None:
            continue
        if product not in target_products:
            continue

        contract_desc = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["contract"]).value or ""
        contract_month = _extract_contract_month(contract_desc)

        participant_id = str(ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["participant_id"]).value or "")
        name_en = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["name_en"]).value or ""
        name_jp = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["name_jp"]).value or ""

        rank_val = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["rank"]).value
        rank = int(rank_val) if rank_val else 0

        vol_val = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["volume"]).value
        volume = _parse_volume_value(vol_val)

        pv = ParticipantVolume(
            trade_date=trade_date,
            product=product,
            contract_month=contract_month,
            participant_id=participant_id,
            participant_name_en=name_en,
            participant_name_jp=name_jp,
            rank=rank,
            volume=volume,
            volume_day=0.0 if is_night else volume,
            volume_night=volume if is_night else 0.0,
        )
        results.append(pv)

    wb.close()
    return results


def _parse_volume_value(val) -> float:
    """Parse volume cell value which may be a number, string like '=21311.0', or None."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("="):
        s = s[1:]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _extract_contract_month(contract_desc: str) -> str:
    """Extract YYMM from contract description.

    'NIKKEI 225 FUT 2603' -> '2603'
    'TOPIX FUT 2603' -> '2603'
    'MINI NK225 FUT 2602' -> '2602'
    """
    if not contract_desc:
        return ""
    parts = contract_desc.strip().split()
    return parts[-1] if parts else ""


def parse_option_volume_excel(content: bytes) -> list[OptionParticipantVolume]:
    """Parse option volume records (NK225E) from a daily volume Excel file.

    Contract format: 'NIKKEI 225 OOP P2602-53250'
      P/C = PUT/CALL, 2602 = YYMM, 53250 = strike price
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    trade_date_str = str(ws.cell(row=5, column=3).value).strip()
    trade_date = datetime.strptime(trade_date_str, "%Y%m%d").date()

    header_text = str(ws.cell(row=2, column=1).value or "")
    is_night = "Night" in header_text

    results = []
    for row_idx in range(config.VOLUME_DATA_START_ROW, ws.max_row + 1):
        product = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["product"]).value
        if product is None or str(product) != "NK225E":
            continue

        contract_desc = str(ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["contract"]).value or "")
        option_type, strike = _parse_option_contract(contract_desc)
        if not option_type:
            continue

        pid = str(ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["participant_id"]).value or "")
        name_en = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["name_en"]).value or ""
        name_jp = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["name_jp"]).value or ""

        rank_val = ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["rank"]).value
        rank = int(rank_val) if rank_val else 0

        vol = _parse_volume_value(
            ws.cell(row=row_idx, column=config.VOLUME_COLUMNS["volume"]).value
        )

        results.append(OptionParticipantVolume(
            trade_date=trade_date,
            option_type=option_type,
            strike_price=strike,
            participant_id=pid,
            participant_name_en=name_en,
            participant_name_jp=name_jp,
            rank=rank,
            volume=vol,
            volume_day=0.0 if is_night else vol,
            volume_night=vol if is_night else 0.0,
        ))

    wb.close()
    return results


def _parse_option_contract(desc: str) -> tuple[str, int]:
    """Parse 'NIKKEI 225 OOP P2602-53250' -> ('PUT', 53250).

    Returns ('', 0) if not parseable.
    """
    # Match P or C followed by YYMM-strike
    m = re.search(r'([PC])(\d{4})-(\d+)', desc)
    if not m:
        return ("", 0)
    opt_char = m.group(1)
    strike = int(m.group(3))
    option_type = "PUT" if opt_char == "P" else "CALL"
    return (option_type, strike)


def merge_option_volume_records(
    *record_lists: list[OptionParticipantVolume],
) -> list[OptionParticipantVolume]:
    """Merge option volume records across sessions."""
    combined: dict[tuple, OptionParticipantVolume] = {}
    for records in record_lists:
        for r in records:
            key = (r.trade_date, r.option_type, r.strike_price, r.participant_id)
            if key in combined:
                existing = combined[key]
                existing.volume += r.volume
                existing.volume_day += r.volume_day
                existing.volume_night += r.volume_night
            else:
                combined[key] = OptionParticipantVolume(
                    trade_date=r.trade_date,
                    option_type=r.option_type,
                    strike_price=r.strike_price,
                    participant_id=r.participant_id,
                    participant_name_en=r.participant_name_en,
                    participant_name_jp=r.participant_name_jp,
                    rank=r.rank,
                    volume=r.volume,
                    volume_day=r.volume_day,
                    volume_night=r.volume_night,
                )
    return list(combined.values())


def merge_volume_records(
    *record_lists: list[ParticipantVolume],
) -> list[ParticipantVolume]:
    """Merge multiple lists of volume records into combined totals.

    Matching on (trade_date, product, contract_month, participant_id).
    Volumes are summed across all input lists.
    Used to combine WholeDay + Night + WholeDayJNet + NightJNet.
    """
    combined: dict[tuple, ParticipantVolume] = {}

    for records in record_lists:
        for r in records:
            key = (r.trade_date, r.product, r.contract_month, r.participant_id)
            if key in combined:
                existing = combined[key]
                existing.volume += r.volume
                existing.volume_day += r.volume_day
                existing.volume_night += r.volume_night
                if not existing.participant_name_en and r.participant_name_en:
                    existing.participant_name_en = r.participant_name_en
                if not existing.participant_name_jp and r.participant_name_jp:
                    existing.participant_name_jp = r.participant_name_jp
            else:
                combined[key] = ParticipantVolume(
                    trade_date=r.trade_date,
                    product=r.product,
                    contract_month=r.contract_month,
                    participant_id=r.participant_id,
                    participant_name_en=r.participant_name_en,
                    participant_name_jp=r.participant_name_jp,
                    rank=r.rank,
                    volume=r.volume,
                    volume_day=r.volume_day,
                    volume_night=r.volume_night,
                )

    return list(combined.values())
