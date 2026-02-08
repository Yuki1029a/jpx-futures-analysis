"""Parse weekly option open interest Excel files.

Excel structure (nk225op_oi_by_tp.xlsx):
  Row 1: Title
  Row 2: Report date  e.g. （ 2026年01月30日現在 ）
  Row 7: C2=プット（2026年02月限月）  C12=コール（2026年02月限月）
  Row 9: C4=（売超参加者）  C7=（買超参加者）  C14=（売超参加者）  C17=（買超参加者）
  Rows 10-24: strike block 1 (15 rows per strike, rank 1-15)
  Rows 25-39: strike block 2
  ...
  PUT side:  C1=rank, C2=strike, C3=short_pid, C4=short_name, C5=short_vol,
             C6=long_pid, C7=long_name, C8=long_vol
  CALL side: C11=rank, C12=strike, C13=short_pid, C14=short_name, C15=short_vol,
             C16=long_pid, C17=long_name, C18=long_vol
"""
from __future__ import annotations

import io
import re
import openpyxl
from datetime import date
from models import OptionParticipantOI


_ROWS_PER_STRIKE = 15

# Column offsets for PUT side (left)
_PUT_COLS = {
    "rank": 1, "strike": 2,
    "short_pid": 3, "short_name": 4, "short_vol": 5,
    "long_pid": 6, "long_name": 7, "long_vol": 8,
}

# Column offsets for CALL side (right)
_CALL_COLS = {
    "rank": 11, "strike": 12,
    "short_pid": 13, "short_name": 14, "short_vol": 15,
    "long_pid": 16, "long_name": 17, "long_vol": 18,
}


def parse_option_oi_excel(content: bytes) -> list[OptionParticipantOI]:
    """Parse option OI Excel and return all PUT/CALL records."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    report_date = _extract_report_date(ws)
    contract_month = _extract_contract_month(ws)
    data_start = _find_data_start(ws)

    results = []
    row = data_start

    while row <= ws.max_row:
        # Parse PUT side of this strike block
        put_strike = ws.cell(row=row, column=_PUT_COLS["strike"]).value
        if put_strike is not None:
            strike_int = int(float(put_strike))
            results.extend(_parse_strike_block(
                ws, row, "PUT", strike_int, report_date, contract_month, _PUT_COLS
            ))

        # Parse CALL side of this strike block
        call_strike = ws.cell(row=row, column=_CALL_COLS["strike"]).value
        if call_strike is not None:
            strike_int = int(float(call_strike))
            results.extend(_parse_strike_block(
                ws, row, "CALL", strike_int, report_date, contract_month, _CALL_COLS
            ))

        row += _ROWS_PER_STRIKE

    wb.close()
    return _consolidate(results)


def _extract_report_date(ws) -> date:
    """Extract report date from row 2: （ 2026年01月30日現在 ）"""
    for row_idx in [2, 3, 1]:
        val = ws.cell(row=row_idx, column=1).value
        if val:
            digits = re.findall(r'\d+', str(val))
            if len(digits) >= 3:
                y, m, d = int(digits[0]), int(digits[1]), int(digits[2])
                if 2000 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
                    return date(y, m, d)
    raise ValueError("Could not extract report date from option OI Excel")


def _extract_contract_month(ws) -> str:
    """Extract contract month from row 7.

    Row 7 C2: 'プット（2026年02月限月）' -> '2602'
    """
    for col in [2, 12]:
        val = ws.cell(row=7, column=col).value
        if val:
            digits = re.findall(r'\d+', str(val))
            if len(digits) >= 2:
                year = digits[0]
                month = digits[1].zfill(2)
                return year[2:] + month
    return ""


def _find_data_start(ws) -> int:
    """Find the first data row (where rank=1 appears in column A)."""
    for row_idx in range(8, min(20, ws.max_row + 1)):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            try:
                if int(val) == 1:
                    return row_idx
            except (ValueError, TypeError):
                pass
    return 10  # fallback


def _parse_strike_block(
    ws, start_row: int, option_type: str, strike: int,
    report_date: date, contract_month: str, cols: dict,
) -> list[OptionParticipantOI]:
    """Parse one strike block (15 rows) for one side (PUT or CALL)."""
    records = []

    for i in range(_ROWS_PER_STRIKE):
        row = start_row + i

        # Short side
        short_pid = ws.cell(row=row, column=cols["short_pid"]).value
        if short_pid:
            short_name = str(ws.cell(row=row, column=cols["short_name"]).value or "")
            short_vol = ws.cell(row=row, column=cols["short_vol"]).value
            records.append(OptionParticipantOI(
                report_date=report_date,
                contract_month=contract_month,
                option_type=option_type,
                strike_price=strike,
                participant_id=str(short_pid),
                participant_name_jp=short_name,
                long_volume=None,
                short_volume=float(short_vol) if short_vol else 0.0,
            ))

        # Long side
        long_pid = ws.cell(row=row, column=cols["long_pid"]).value
        if long_pid:
            long_name = str(ws.cell(row=row, column=cols["long_name"]).value or "")
            long_vol = ws.cell(row=row, column=cols["long_vol"]).value
            records.append(OptionParticipantOI(
                report_date=report_date,
                contract_month=contract_month,
                option_type=option_type,
                strike_price=strike,
                participant_id=str(long_pid),
                participant_name_jp=long_name,
                long_volume=float(long_vol) if long_vol else 0.0,
                short_volume=None,
            ))

    return records


def _consolidate(records: list[OptionParticipantOI]) -> list[OptionParticipantOI]:
    """Merge long/short into single records per (contract_month, option_type, strike, pid)."""
    key_map: dict[tuple, OptionParticipantOI] = {}
    for rec in records:
        key = (rec.contract_month, rec.option_type, rec.strike_price, rec.participant_id)
        if key in key_map:
            existing = key_map[key]
            if rec.long_volume is not None:
                existing.long_volume = rec.long_volume
            if rec.short_volume is not None:
                existing.short_volume = rec.short_volume
        else:
            key_map[key] = rec
    return list(key_map.values())
