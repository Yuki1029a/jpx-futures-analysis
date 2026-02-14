"""Analyze the structure of option OI Excel file."""

import openpyxl
from pathlib import Path

file_path = Path("cache/oi/20260130_nk225op_oi_by_tp.xlsx")

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print("\n" + "="*80)
print("First 30 rows:")
print("="*80)

for row_idx in range(1, min(31, ws.max_row + 1)):
    values = []
    for col_idx in range(1, min(11, ws.max_column + 1)):
        val = ws.cell(row=row_idx, column=col_idx).value
        val_str = str(val) if val is not None else ""
        if len(val_str) > 15:
            val_str = val_str[:12] + "..."
        values.append(val_str)
    print(f"Row {row_idx:2d}: {' | '.join(values)}")

print("\n" + "="*80)
print("Searching for 'PUT' keyword:")
print("="*80)

for row_idx in range(1, min(100, ws.max_row + 1)):
    for col_idx in range(1, min(15, ws.max_column + 1)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and 'PUT' in str(val).upper():
            print(f"Row {row_idx}, Col {col_idx}: {val}")

print("\n" + "="*80)
print("Searching for 'CALL' keyword:")
print("="*80)

for row_idx in range(1, min(100, ws.max_row + 1)):
    for col_idx in range(1, min(15, ws.max_column + 1)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and 'CALL' in str(val).upper():
            print(f"Row {row_idx}, Col {col_idx}: {val}")

wb.close()
